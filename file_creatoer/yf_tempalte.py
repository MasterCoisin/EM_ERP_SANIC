# -*- coding: UTF-8 -*-
'''
@Project     ：EM_ERP_BACKEN 
@File        ：yf_tempalte.py
@IDE         ：PyCharm 
-------------------------------------
@Author      ：Coisin
@QQ          ：2849068933
@PHONE       ：17350199092
@Description ：运费模板生成
@Date        ：2024-07-12 9:22 
-------------------------------------
'''
import io
import time
from uuid import uuid1
from openpyxl.drawing.image import Image
import os
from openpyxl import load_workbook
from sanic import response

column_width = 10
row_height = 80  # 设置行高，该设置的行高与excel文件中设置的行高值是一样的

BASE_PATH = os.getcwd()
print(BASE_PATH)
def form_tempalte_pc(packing_info,ean_info):
    wb = load_workbook(r'C:/myprograms/爬虫/EM/利润计算/EM_ERP_BAKEN_SANIC/excel_template/yfmb_pc.xlsx')
    sheet = wb.active
    row = 20
    total_box = 0
    for box_index, box_data in enumerate(packing_info):
        box_index += 1
        total_box += 1
        msku_in_box = box_data.get('mskuInBox', {})
        for ean,count in msku_in_box.items():
            if count > 0:
                ean_data = ean_info.get(ean,{})
                images = ean_data.get("images",None)
                invoice_info = ean_data.get("invoice_info",{})
                em_attribute = ean_data.get("em_attribute",{})
                length_width_height_weight = ean_data.get("length_width_height_weight",{})
                flag = ean_data.get("flag",{})
                if flag.get("isElectrified", False):
                    sheet.cell(row=2, column=6).value = "是"
                if flag.get("isMagnetized", False):
                    sheet.cell(row=3, column=6).value = "是"
                if flag.get("isLiquid", False):
                    sheet.cell(row=5, column=6).value = "是"
                if flag.get("isPowder", False):
                    sheet.cell(row=4, column=6).value = "是"
                sheet.cell(row=row, column=1).value = box_index
                sheet.cell(row=row, column=2).value = invoice_info.get('productNameZh', None)
                sheet.cell(row=row, column=3).value = invoice_info.get('productNameEn', None)
                sheet.cell(row=row, column=4).value = invoice_info.get('product_customs_code', None)
                sheet.cell(row=row, column=5).value = count
                price = float(invoice_info.get('productDeclarationUnitPrice', None)) if invoice_info.get(
                    'productDeclarationUnitPrice', None) is not None else None
                sheet.cell(row=row, column=6).value = price
                sheet.cell(row=row, column=7).value = price * count if invoice_info.get('productDeclarationUnitPrice', None) is not None else None
                sheet.cell(row=row, column=8).value = invoice_info.get('productMaterial', None)
                sheet.cell(row=row, column=9).value = invoice_info.get('productUsage', None)
                sheet.cell(row=row, column=10).value = length_width_height_weight.get('G', None)
                sheet.cell(row=row, column=11).value = length_width_height_weight.get('G', None)
                sheet.cell(row=row, column=16).value = f"https://www.emag.ro/-/pd/{em_attribute.get('pnk',{}).get('ro',None)}/"
                sheet.cell(row=row, column=17).value = f"无"
                sheet.cell(row=row, column=18).value = f"无"
                # 下面代码中的[]括号中可以输入'D'或者'd'
                sheet.column_dimensions['O'].width = column_width  # 修改列D的列宽
                sheet.row_dimensions[row].height = row_height  # 修改行3的行高
                # 把二进制数据转为类文件对象
                if images:
                    print(images[0])
                    img_stream = io.BytesIO(images[0])
                    img = Image(img_stream)  # 调用图像函数
                    newSize = (90, 90)
                    img.width, img.height = newSize  # 这两个属性分别是对应添加图片的宽高
                    sheet.add_image(img, 'O' + str(row))  # 向d列中的单元格内指定添加图片
                row += 1
    sheet.cell(row=1, column=6).value = total_box
    # 创建一个内存中的二进制流
    output = io.BytesIO()
    # 将 Excel 文件保存到流中
    wb.save(output)
    # 将文件指针移动到流的开头
    output.seek(0)
    # 返回文件响应
    return response.raw(
        output.read(),
        headers={
            'Content-Disposition': f'attachment; filename=鹏程运费模板-{time.time()}.xlsx',
            'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        }
    )


def form_tempalte_ynd(packing_info,ean_info):
    wb = load_workbook(os.path.join(BASE_PATH,r'/excel_template/yfmb_ynd.xlsx'))
    sheet = wb.active
    row = 17
    total_box = 0
    for box_index, box_data in enumerate(packing_info):
        box_index += 1
        total_box+=1
        msku_in_box = box_data.get('mskuInBox', {})
        for ean,count in msku_in_box.items():
            if count > 0:
                ean_data = ean_info.get(ean, {})
                msku = ean_data.get("msku", None)
                images = ean_data.get("images", None)
                invoice_info = ean_data.get("invoice_info", {})
                em_attribute = ean_data.get("em_attribute", {})
                length_width_height_weight = ean_data.get("length_width_height_weight", {})
                flag = ean_data.get("flag", {})
                if flag.get("isElectrified",False):
                    sheet.cell(row=1, column=6).value = "是"
                if flag.get("isMagnetized",False):
                    sheet.cell(row=2, column=6).value = "是"
                if flag.get("isLiquid",False):
                    sheet.cell(row=3, column=6).value = "是"
                if flag.get("isPowder",False):
                    sheet.cell(row=4, column=6).value = "是"
                if flag.get("isDangerous",False):
                    sheet.cell(row=5, column=6).value = "是"
                #
                sheet.cell(row=row, column=1).value = box_index
                sheet.cell(row=row, column=6).value = msku
                sheet.cell(row=row, column=8).value = invoice_info.get('productNameZh', None)
                sheet.cell(row=row, column=7).value = invoice_info.get('productNameEn', None)
                sheet.cell(row=row, column=13).value = invoice_info.get('productCustomsCode', None)
                sheet.cell(row=row, column=10).value = count
                price = float(invoice_info.get('productDeclarationUnitPrice', None)) if invoice_info.get(
                    'productDeclarationUnitPrice', None) is not None else None
                sheet.cell(row=row, column=9).value = price
                sheet.cell(row=row, column=11).value = invoice_info.get('productMaterial', None)
                sheet.cell(row=row, column=12).value = invoice_info.get('productUsage', None)
                sheet.cell(row=row, column=16).value = f"https://www.emag.ro/-/pd/{em_attribute.get('pnk',{}).get('ro',None)}/"
                sheet.cell(row=row, column=14).value = f"无"
                sheet.cell(row=row, column=15).value = f"无"
                # 把二进制数据转为类文件对象
                if images:
                    img_stream = io.BytesIO(images[0])
                    img = Image(img_stream)  # 调用图像函数
                    newSize = (30, 30)
                    img.width, img.height = newSize  # 这两个属性分别是对应添加图片的宽高
                    sheet.add_image(img, 'Q' + str(row))  # 向d列中的单元格内指定添加图片
                row += 1
    sheet.cell(row=15, column=2).value = total_box
    # 创建一个内存中的二进制流
    output = io.BytesIO()
    # 将 Excel 文件保存到流中
    wb.save(output)
    # 将文件指针移动到流的开头
    output.seek(0)
    # 返回文件响应
    return response.raw(
        output.read(),
        headers={
            'Content-Disposition': f'attachment; filename=伊诺达运费模板-{time.time()}.xlsx',
            'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        }
    )
