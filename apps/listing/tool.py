# -*- coding: UTF-8 -*-
'''
@Project     ：EM_ERP_BAKEN_SANIC 
@File        ：tool.py
@IDE         ：PyCharm 
-------------------------------------
@Author      ：Coisin
@QQ          ：2849068933
@PHONE       ：17350199092
@Description ：
@Date        ：2025-02-11 16:53 
-------------------------------------
'''
import io
import json
import time

import aiohttp
from loguru import logger
from motor.motor_asyncio import AsyncIOMotorCollection
from sanic import Sanic, response

from apps.em_category.tool import get_em_category_name
from apps.images.tool import image_id_to_url, get_image_emUrl, get_image_source_img_file_by_url_or_source_id
from config.constant import APP_NAME, VAT, EXCHANGE
from models.listing import Listing
from utils.cal_fbe_fee import CalFbeFee
from utils.common import transfer_date, parse_str_to_float, get_html_image_detail, args_valid
from utils.genius_fee_cal import genius_fee_cal_ro, genius_fee_cal_hu, genius_fee_cal_bg
from openpyxl import load_workbook

app = Sanic.get_app(APP_NAME)
fbe_fee_cal_obj = CalFbeFee()


async def get_listing_by_ean(ean):
    collection: AsyncIOMotorCollection = app.ctx.mongo[Listing.collection_name]
    result = await collection.find_one({"ean": ean})
    return result


async def get_listing_shipping_info(ean):
    collection: AsyncIOMotorCollection = app.ctx.mongo[Listing.collection_name]
    obj = await collection.find_one({"ean": ean})
    res = {}
    res["listing_name"] = obj.get("listingName", None)
    res["images"] = obj.get("images", [])
    res["ean"] = obj.get("ean", None)
    res["flag"] = obj.get("flag", {})
    logistics_attributes = obj.get("logisticsAttributes", None)
    res["length_width_height_weight"] = {"L": parse_str_to_float(logistics_attributes.get("length", None)),
                                         "W": parse_str_to_float(logistics_attributes.get("width", None)),
                                         "H": parse_str_to_float(logistics_attributes.get("height", None)),
                                         "G": parse_str_to_float(logistics_attributes.get("weight", None),
                                                                 1000)}
    res['invoice_info'] = obj.get("invoiceInfo", None)
    res['em_attribute'] = obj.get("emAttribute", None)
    try:
        res["images"] = [await get_image_source_img_file_by_url_or_source_id(i) for i in res["images"][:1]]
    except Exception as e:
        logger.exception(e)
        pass
    res["msku"] = obj.get("msku", None)
    return res


async def get_data_base_info_by_eans(eans, campany_id):
    collection: AsyncIOMotorCollection = app.ctx.mongo[Listing.collection_name]
    result = await collection.find({"ean": {"$in": eans}, "campanyId": campany_id},
                                   {"_id": 0, "ean": 1, "images": 1, "listingName": 1}).to_list()
    return {i["ean"]: i for i in result}


async def get_weighing_info_by_eans(eans, campany_id):
    collection: AsyncIOMotorCollection = app.ctx.mongo[Listing.collection_name]
    result = await collection.find({"ean": {"$in": eans}, "campanyId": campany_id},
                                   {"_id": 0, "ean": 1, "logisticsAttributes": 1}).to_list()
    return {i["ean"]: i.get("logisticsAttributes", {}).get("weighing", None) for i in result}


async def get_ean_and_images_by_id(campany_id, shop, product_ids, country):
    collection: AsyncIOMotorCollection = app.ctx.mongo[Listing.collection_name]
    result = await collection.find(
        {f"emAttribute.id.{country}": {"$in": product_ids}, "campanyId": campany_id, "shop": shop},
        {"_id": 0, "ean": 1, "images": 1, "logisticsAttributes": 1, "emAttribute.id": 1, "emSaleInfo.stock": 1,
         "emAttribute.commission": 1}).to_list()
    return {i["emAttribute"]["id"][country]: {"ean": i["ean"], "images": i["images"],
                                              "stock": i["emSaleInfo"]["stock"][country],
                                              "commission": i["emAttribute"]["commission"][country],
                                              "lengthWidthHeightWeight": {
                                                  "L": i.get("logisticsAttributes", {}).get("length", None),
                                                  "W": i.get("logisticsAttributes", {}).get("width", None),
                                                  "H": i.get("logisticsAttributes", {}).get("height", None),
                                                  "G": parse_str_to_float(
                                                      i.get("logisticsAttributes", {}).get("weight", None), 1000)}
                                              } for i in result}


async def get_stock_info_by_eans(eans, campany_id):
    collection: AsyncIOMotorCollection = app.ctx.mongo[Listing.collection_name]
    result = await collection.find({"ean": {"$in": eans}, "campanyId": campany_id},
                                   {"_id": 0, "ean": 1, "images": 1, "emSaleInfo.stock": 1}).to_list()
    return {i["ean"]: {"images": i["images"], "stock": i["emSaleInfo"]["stock"]["ro"]} for i in result}


async def update_listing_by_ean(ean, data):
    await app.ctx.mongo[Listing.collection_name].update_one({"ean": ean},
                                                            {"$set": data})


async def parse_to_list(org_data, ean_wh_data):
    # 合并库存信息
    org_data["whData"] = ean_wh_data
    # 图片
    org_data["images"] = [image_id_to_url(img_info) for img_info in org_data["images"]]
    # 时间转化
    if org_data.get("first_create_date"):
        org_data["first_create_date"] = transfer_date(org_data["first_create_date"], False)
    else:
        org_data["first_create_date"] = None
    # 尺寸数据是否完成
    logistics_attributes = org_data.get("logisticsAttributes", {})
    org_data["completedLengthWidthHeightWeight"] = (
            logistics_attributes.get("length", None) not in [None, 0, "0"] and logistics_attributes.get(
        "width", None) not in [None, 0, "0"] and logistics_attributes.get("height",
                                                                          None) not in [None, 0,
                                                                                        "0"] and logistics_attributes.get(
        "weight", None) not in [None, 0, "0"] and logistics_attributes.get("length",
                                                                           None) != "" and logistics_attributes.get(
        "width", None) != "" and logistics_attributes.get("height", None) != "" and logistics_attributes.get(
        "weight", None) != "")
    org_data["lengthWidthHeightWeight"] = {"L": logistics_attributes.get("length", None),
                                           "W": logistics_attributes.get("width", None),
                                           "H": logistics_attributes.get("height", None),
                                           "G": parse_str_to_float(logistics_attributes.get("weight", None), 1000)}
    # 计算fbe数据
    org_data["fbeFee"] = await fbe_fee_cal_obj.cal_fbe_fee(
        org_data["lengthWidthHeightWeight"]["G"] * 1000 if org_data["lengthWidthHeightWeight"]["G"] else
        org_data["lengthWidthHeightWeight"]["G"],
        org_data["lengthWidthHeightWeight"]["L"] if org_data["lengthWidthHeightWeight"]["L"] else
        org_data["lengthWidthHeightWeight"]["L"],
        org_data["lengthWidthHeightWeight"]["W"] if org_data["lengthWidthHeightWeight"]["W"] else
        org_data["lengthWidthHeightWeight"]["W"],
        org_data["lengthWidthHeightWeight"]["H"] if org_data["lengthWidthHeightWeight"]["H"] else
        org_data["lengthWidthHeightWeight"]["H"])
    #############################
    # 利润计算
    profit = {
        "profit": {
            "ro": None,
            "bg": None,
            "hu": None
        },  # 利润率
        "profitRmb": {
            "ro": None,
            "bg": None,
            "hu": None
        },  # 利润
        "soldPriceWithVat": {
            "ro": None,
            "bg": None,
            "hu": None
        },  # 含税售价
        "soldPriceWithoutVat": {
            "ro": None,
            "bg": None,
            "hu": None
        },  # 含税售价
        "geniusFee": {
            "ro": 0,
            "bg": 0,
            "hu": 0
        },
        "commissionFee": {
            "ro": None,
            "bg": None,
            "hu": None
        }
    }
    # 成本项
    msku_avg_buy_and_head_fee = org_data["whData"].get("mskuAvgBuyAndHeadFee", None)  # 采运成本RMB
    order_fee_rmb = org_data["fbeFee"].get("orderFeeRmb", None)
    # 售价
    sale_price_without_vat_ro = org_data["emAttribute"].get("salePrice", {}).get("ro", None)
    profit["soldPriceWithoutVat"]["ro"] = sale_price_without_vat_ro
    profit["soldPriceWithVat"]["ro"] = sale_price_without_vat_ro * (
            1 + VAT["ro"]) if sale_price_without_vat_ro is not None else None
    sale_price_without_vat_bg = org_data["emAttribute"].get("salePrice", {}).get("bg", None)
    profit["soldPriceWithoutVat"]["bg"] = sale_price_without_vat_bg
    profit["soldPriceWithVat"]["bg"] = sale_price_without_vat_bg * (
            1 + VAT["bg"]) if sale_price_without_vat_bg is not None else None
    sale_price_without_vat_hu = org_data["emAttribute"].get("salePrice", {}).get("hu", None)
    profit["soldPriceWithoutVat"]["hu"] = sale_price_without_vat_hu
    profit["soldPriceWithVat"]["hu"] = sale_price_without_vat_hu * (
            1 + VAT["hu"]) if sale_price_without_vat_hu is not None else None
    # 计算geniusFee
    if profit["soldPriceWithVat"]["ro"] is not None and org_data["emSaleInfo"]["geniusEligibility"]["ro"] == 1:
        profit["geniusFee"]["ro"] = genius_fee_cal_ro(profit["soldPriceWithVat"]["ro"])
    if profit["soldPriceWithVat"]["bg"] is not None and org_data["emSaleInfo"]["geniusEligibility"]["bg"] == 1:
        profit["geniusFee"]["bg"] = genius_fee_cal_bg(profit["soldPriceWithVat"]["bg"])
    if profit["soldPriceWithVat"]["hu"] is not None and org_data["emSaleInfo"]["geniusEligibility"]["hu"] == 1:
        profit["geniusFee"]["hu"] = genius_fee_cal_hu(profit["soldPriceWithVat"]["hu"])
    # 佣金
    profit["commissionFee"]["ro"] = org_data["emAttribute"]["commission"][
                                        "ro"] * sale_price_without_vat_ro * EXCHANGE[
                                        "ro"] / 100.0 if (
            sale_price_without_vat_ro is not None and org_data["emAttribute"]["commission"][
        "ro"] is not None) else None
    profit["commissionFee"]["bg"] = org_data["emAttribute"]["commission"][
                                        "bg"] * sale_price_without_vat_bg * EXCHANGE[
                                        "bg"] / 100.0 if (
            sale_price_without_vat_bg is not None and org_data["emAttribute"]["commission"][
        "bg"] is not None) else None
    profit["commissionFee"]["hu"] = org_data["emAttribute"]["commission"][
                                        "hu"] * sale_price_without_vat_hu * EXCHANGE[
                                        "hu"] / 100.0 if (
            sale_price_without_vat_hu is not None and org_data["emAttribute"]["commission"][
        "hu"] is not None) else None
    # 计算利润
    for country in ["ro", "bg", "hu"]:
        if args_valid([msku_avg_buy_and_head_fee, order_fee_rmb, profit["soldPriceWithoutVat"][country],
                       profit["soldPriceWithVat"][country], profit["geniusFee"][country],
                       profit["commissionFee"][country]]):
            profit["profitRmb"][country] = profit["soldPriceWithoutVat"][
                                               country] * EXCHANGE[
                                               country] - msku_avg_buy_and_head_fee - order_fee_rmb - \
                                           profit["geniusFee"][country] - profit["commissionFee"][country]
            profit["profit"][
                country] = f'{round(100.0 * profit["profitRmb"][country] / (profit["soldPriceWithVat"][country] * EXCHANGE[country]), 2)} %'
        profit["profitRmb"][country] = f'{round(profit["profitRmb"][country], 2)}(￥)' if profit["profitRmb"][
            country] else profit["profitRmb"][country]
        profit["commissionFee"][country] = f'{round(profit["commissionFee"][country], 2)}(￥)' if \
            profit["commissionFee"][country] else profit["commissionFee"][country]
    org_data["profit"] = profit
    #############################
    # 装箱发票信息是否完成
    invoice_info = org_data.get("invoiceInfo", {})
    if invoice_info:
        org_data["invoiceInfoCompleted"] = True
        for k, v in invoice_info.items():
            if v is None or v == "":
                org_data["invoiceInfoCompleted"] = False
    else:
        org_data["invoiceInfoCompleted"] = False
    # 是否满足上传条件
    org_data["couldUpdate"] = True
    em_attribute = org_data.get("emAttribute", {})
    if em_attribute.get("brand", {}).get("ro", None) and em_attribute.get("category", {}).get("ro",
                                                                                              None) and em_attribute.get(
        "html",
        {}).get("ro", None) and em_attribute.get(
        "title", {}).get("ro", None) and em_attribute.get("salePrice", {}).get("ro", None):
        pass
    else:
        org_data["couldUpdate"] = False

    # 说明书
    org_data["shuomingLanguages"] = ",".join([i["language"] for i in org_data.get("gprs", {}).get("details", {})])

    #
    # em_sold_price_without_vat = em_attribute.get("saleprice", None)
    # em_sold_price_without_vat = float(
    #     em_sold_price_without_vat) if em_sold_price_without_vat else em_sold_price_without_vat
    # sold_price = org_data.get("productProfitCalData", {}).get("soldPriceWithVat",
    #                                                           em_sold_price_without_vat * (
    #                                                                   1 + VAT) if em_sold_price_without_vat else None)
    # org_data["profitCalResult"]["soldPriceWithoutVat"] = em_sold_price_without_vat
    # org_data["profitCalResult"]["soldPriceWithVat"] = sold_price
    # 利润计算

    # 删除字段
    # del org_data["gprs"]
    return org_data


def parse_to_packing_data(org_data):
    result = {"ean": org_data["ean"], "images": [image_id_to_url(img_info) for img_info in org_data["images"]],
              "listingName": org_data["listingName"], "flag": org_data["flag"],
              "id": org_data.get("emAttribute", {}).get("id", {}).get("ro", None)}
    # 图片
    # 尺寸数据是否完成
    logistics_attributes = org_data.get("logisticsAttributes", {})
    completedLengthWidthHeightWeight = (
            logistics_attributes.get("length", None) not in [None, 0, "0"] and logistics_attributes.get(
        "width", None) not in [None, 0, "0"] and logistics_attributes.get("height",
                                                                          None) not in [None, 0,
                                                                                        "0"] and logistics_attributes.get(
        "weight", None) not in [None, 0, "0"] and logistics_attributes.get("length",
                                                                           None) != "" and logistics_attributes.get(
        "width", None) != "" and logistics_attributes.get("height", None) != "" and logistics_attributes.get(
        "weight", None) != "")
    result["length"] = logistics_attributes.get("length", None) / 10.0 if logistics_attributes.get("length",
                                                                                                   None) else logistics_attributes.get(
        "length", None)
    result["width"] = logistics_attributes.get("width", None) / 10.0 if logistics_attributes.get("width",
                                                                                                 None) else logistics_attributes.get(
        "width", None)
    result["height"] = logistics_attributes.get("height", None) / 10.0 if logistics_attributes.get("height",
                                                                                                   None) else logistics_attributes.get(
        "height", None)
    result["weight"] = logistics_attributes.get("weight", None) / 1000.0 if logistics_attributes.get("weight",
                                                                                                     None) else logistics_attributes.get(
        "weight", None)

    # 装箱发票信息是否完成
    invoice_info = org_data.get("invoiceInfo", {})
    invoiceInfoCompleted = True
    if invoice_info:
        for k, v in invoice_info.items():
            if v is None or v == "":
                invoiceInfoCompleted = False
    else:
        invoiceInfoCompleted = False
    pnk = org_data.get("emAttribute", {}).get("pnk", {}).get("ro", None)
    return completedLengthWidthHeightWeight, invoiceInfoCompleted, pnk is not None, result


async def get_packing_data_by_ean(ean, campany_id):
    collection: AsyncIOMotorCollection = app.ctx.mongo[Listing.collection_name]
    result = await collection.find_one({"ean": ean, "campanyId": campany_id, "deleted": False})
    if not result:
        result = await collection.find_one({"msku": ean, "campanyId": campany_id, "deleted": False})
    if result:
        completedLengthWidthHeightWeight, invoiceInfoCompleted, has_pnk, result = parse_to_packing_data(result)
        if not has_pnk:
            return False, "产品未完成审核"
        if not completedLengthWidthHeightWeight:
            return False, "请完善产品尺寸信息！"
        if not invoiceInfoCompleted:
            return False, "请完善产品发票信息！"
        return True, result
    else:
        return False, "未找到产品"


async def get_sku_data_by_ean(ean, campany_id):
    collection: AsyncIOMotorCollection = app.ctx.mongo[Listing.collection_name]
    result = await collection.find_one({"ean": ean, "campanyId": campany_id, "deleted": False}, {"skuList": 1})
    if result:
        return True, result.get("skuList", [])
    else:
        return False, "未找到产品"


async def save_expain_to_server(ean, html):
    url = "https://walawaladocs.de/api/pull_data/"
    payload = json.dumps({
        "ean": ean,
        "html": html
    })
    headers = {
        'Content-Type': 'application/json'
    }
    try:
        async with aiohttp.ClientSession() as session:
            async with session.post(url, headers=headers, data=payload, timeout=5) as response:
                if response.status == 200:
                    res = await response.json()
                    return True
                raise Exception(f"Request failed with status code {response.status}")
    except Exception as e:
        return False


async def create_em_product_excel(eans, campany_id):
    collection: AsyncIOMotorCollection = app.ctx.mongo[Listing.collection_name]
    objs = await collection.find({"ean": {"$in": eans}, "campanyId": campany_id}).to_list()
    if not objs:
        return None
    result = []
    col = ["Product code", "Product ID", "Product EAN", "Category ID", "Category name", "Sale price", "RRP", "VAT",
           "Currency", "Stock", "Name", "Brand", "Description", "URL main image", "Other image URL 1",
           "Other image URL 2", "Other image URL 3", "Other image URL 4", "Other image URL 5"]
    for obj in objs:
        msku = obj.get("msku", None)
        ean = obj.get("ean", None)
        em_attribute = obj.get("emAttribute", {})
        brand = em_attribute.get("brand", {}).get("ro", "OEM")
        category = em_attribute.get("category", {}).get("ro", None)
        category_name = await get_em_category_name(category)
        sale_price = em_attribute.get("salePrice", {}).get("ro", None)
        PRP = round(sale_price * 2, 2) if sale_price else None
        title = em_attribute.get("title", {}).get("ro", None)
        html = em_attribute.get("html", {}).get("ro", None)
        images = obj.get("images", [])
        base_images = [None, None, None, None, None, None]
        for idx, i in enumerate(images[:6]):
            if i:
                if "http" not in i:
                    j = await get_image_emUrl(i)
                    if j:
                        base_images[idx] = j
                    else:
                        base_images[idx] = f"https://www.fzwala.com/api/images/get/{i}"
                else:
                    base_images[idx] = i
        result.append([
                          msku,
                          int(ean),
                          category,
                          category_name,
                          ean,
                          sale_price,
                          PRP,
                          0.19,
                          0,
                          "RON",
                          0,
                          10,
                          10000,
                          title,
                          brand,
                          html
                      ] + base_images + [None, "English (en_GB)"])  # "Romanian (ro_RO) English (en_GB)"
    file = create_en_new_products_translation_xlsx(result)
    return file


def create_en_new_products_translation_xlsx(data):
    wb = load_workbook(r'excel_template/en_new_products_translation.xlsx')
    sheet = wb.active
    for row, row_data in enumerate(data):
        for col, col_data in enumerate(row_data):
            sheet.cell(row=6 + row, column=1 + col).value = col_data
    # 创建一个内存中的二进制流
    output = io.BytesIO()
    # 将 Excel 文件保存到流中
    wb.save(output)
    # 将文件指针移动到流的开头
    output.seek(0)
    # 返回文件响应
    return response.raw(
        output.read(),
        headers={
            'Content-Disposition': f'attachment; filename=en_new_products_translation-{time.time()}.xlsx',
            'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        }
    )
