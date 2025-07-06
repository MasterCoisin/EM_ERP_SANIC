# -*- coding: UTF-8 -*-
'''
@Project     ：EM_ERP_BAKEN_SANIC 
@File        ：tool.py
@IDE         ：PyCharm 
-------------------------------------
@Author      ：Coisin
@QQ          ：2849068933
@PHONE       ：17350199092
@Description ：
@Date        ：2025-04-23 16:46 
-------------------------------------
'''
import datetime
import io
import time

from sanic import response

from loguru import logger
from motor.motor_asyncio import AsyncIOMotorCollection
from openpyxl import load_workbook
from sanic import Blueprint, json, Request, Sanic

from apps.sku_inventory_detail.tool import change_sku_status_for_shipments_order
from models.em_reception import EmReception
from config.constant import APP_NAME
from models.shipments_order import ShipmentsOrder
from utils.common import datatime_to_timesmap, get_shipment_order_id

app = Sanic.get_app(APP_NAME)


async def get_em_reception_by_em_reception_id(em_reception_id, campany_id):
    collection: AsyncIOMotorCollection = app.ctx.mongo[EmReception.collection_name]
    result = await collection.find_one({"emReceptionId": em_reception_id, "campanyId": campany_id, "deleted": False},
                                       {"_id": 0})
    if result:
        result = datatime_to_timesmap(result)
        return True, result
    else:
        return False, "未找到面单"


async def get_em_reception_emid_by_em_reception_ids(em_reception_ids, campany_id):
    collection: AsyncIOMotorCollection = app.ctx.mongo[EmReception.collection_name]
    result = await collection.find({"emReceptionId": {"$in": em_reception_ids}, "campanyId": campany_id},
                                   {"_id": 0, "receptionId": 1,"emReceptionId":1}).to_list()
    return {i["emReceptionId"]: i.get("receptionId", None) for i in result}


async def form_em_reception_excel(em_reception_id, receptionId, company_info, rep, eanInfo, campany_id):
    collection: AsyncIOMotorCollection = app.ctx.mongo[EmReception.collection_name]
    await collection.update_one({"emReceptionId": em_reception_id, "campanyId": campany_id, "deleted": False},
                                {"$set": {"receptionId": receptionId}})
    wb = load_workbook(r'excel_template/reception.xlsx')
    sheet = wb.active
    for row, row_data in enumerate(eanInfo):
        ean = row_data.get("ean", None)
        count = row_data.get("count", None)
        length = row_data.get("length", None)
        width = row_data.get("width", None)
        height = row_data.get("height", None)
        weight = row_data.get("weight", None)
        id_ = row_data.get("id", None)
        sheet.cell(row=6 + row, column=1).value = receptionId
        sheet.cell(row=6 + row, column=8).value = str(id_) if id_ is not None else None
        sheet.cell(row=6 + row, column=9).value = str(ean) if ean is not None else None
        sheet.cell(row=6 + row, column=10).value = company_info.get("manufacturer", None)
        sheet.cell(row=6 + row, column=11).value = company_info.get("address", None)
        sheet.cell(row=6 + row, column=12).value = company_info.get("manufacturer_email", None)
        sheet.cell(row=6 + row, column=13).value = rep.get("name", None)
        sheet.cell(row=6 + row, column=14).value = rep.get("address", None)
        sheet.cell(row=6 + row, column=15).value = rep.get("email", None)
        sheet.cell(row=6 + row, column=16).value = "1"
        sheet.cell(row=6 + row, column=17).value = str(count)
        sheet.cell(row=6 + row, column=18).value = "1"
        sheet.cell(row=6 + row, column=20).value = "1"
        sheet.cell(row=6 + row, column=24).value = str(length) if length is not None else None
        sheet.cell(row=6 + row, column=25).value = str(width) if width is not None else None
        sheet.cell(row=6 + row, column=26).value = str(height) if height is not None else None
        sheet.cell(row=6 + row, column=27).value = str(weight) if weight is not None else None
    try:
        # 创建一个内存中的二进制流
        output = io.BytesIO()
        # 将 Excel 文件保存到流中
        wb.save(output)
        # 将文件指针移动到流的开头
        output.seek(0)
        # 返回文件响应
        return response.raw(
            output.read(),
            headers={
                'Content-Disposition': f'attachment; filename=reception-{time.time()}.xlsx',
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            }
        )
    except Exception as e:
        print(f"Error: {e}")
        return response.json({"error": str(e)}, status=500)
    finally:
        # 关闭流
        try:
            output.close()
        except:
            pass


async def change_ean_status(em_reception_id, campany_id, eans, uploadStatus, ):
    collection: AsyncIOMotorCollection = app.ctx.mongo[EmReception.collection_name]
    result = await collection.find_one({"emReceptionId": em_reception_id, "campanyId": campany_id, "deleted": False},
                                       {"_id": 0})
    if result:
        eanInfo = result.get("eanInfo", [])
        for i in eanInfo:
            if i["ean"] in eans:
                i["uploadStatus"] = uploadStatus
        await collection.update_one({"emReceptionId": em_reception_id, "campanyId": campany_id, "deleted": False},
                                    {"$set": {"eanInfo": eanInfo}})
        return True, "OK"
    else:
        return False, "未找到面单"


async def create_shipments_order(emReceptionId, campany_id, domesticLogistics, internationalLogistics, sendToWhId):
    collection: AsyncIOMotorCollection = app.ctx.mongo[EmReception.collection_name]
    collection_shipments_order: AsyncIOMotorCollection = app.ctx.mongo[ShipmentsOrder.collection_name]
    result = await collection.find_one({"emReceptionId": emReceptionId, "campanyId": campany_id, "deleted": False},
                                       {"_id": 0})
    if result:
        packingOrderId = result.get("packingOrderId", None)
        shipmentsOrderId = get_shipment_order_id()
        eanInfo = result.get("eanInfo", [])
        sendFromWhId = result.get("sendFromWhId", None)
        shop = result.get("shop", None)
        # 更新sku状态
        await change_sku_status_for_shipments_order(campany_id, packingOrderId, shipmentsOrderId, eanInfo, sendFromWhId,
                                                    sendToWhId)
        #
        await collection_shipments_order.insert_one({
            "campanyId": campany_id,
            "shop": shop,
            "shipmentsOrderId": shipmentsOrderId,
            "packingOrderId": packingOrderId,
            "emReceptionId": emReceptionId,
            "status": 1,
            "sendFromWhId": sendFromWhId,
            "sendToWhId": sendToWhId,
            "eanInfo": [{"ean": ean["ean"], "count": ean["count"]} for ean in eanInfo],
            "domesticLogisticsInfo": {
                "lpId": domesticLogistics,
                "price": None,
                "priceSure": False,
                "status": 0,
                "t": [
                    None,
                    None,
                    None,
                    None
                ],
                "tNo": "",
                "tNoSure": False
            },
            "internationalLogisticsInfo": {
                "lpId": internationalLogistics,
                "price": None,
                "priceSure": False,
                "status": 0,
                "t": [
                    None,
                    None,
                    None,
                    None
                ],
                "tNo": "",
                "tNoSure": False
            },
            "createTime": datetime.datetime.now(),
            "updateTime": datetime.datetime.now(),
            "deleted": False,
        })
        await collection.update_one({"emReceptionId": emReceptionId, "campanyId": campany_id, "deleted": False},
                                    {"$set": {"sendToWhId": sendToWhId, "shipmentsOrderId": shipmentsOrderId,
                                              "hasSend": True}})
        return True, shipmentsOrderId
    else:
        return False, "未找到面单"
